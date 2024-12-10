import openpyxl

class Excel_file:
    def read_excel_data(self,excel_file,sheet_name,row,col):
        excel=openpyxl.load_workbook(excel_file)
        sheet=excel[sheet_name]
        return sheet.cell(row=row,column=col).value

    def max_row_count(self,excel_file,sheet_name):
        excel=openpyxl.load_workbook(excel_file)
        sheet=excel[sheet_name]
        return sheet.max_row

    def write_excel_data(self,excel_file,sheet_name,row,col,data):
        excel=openpyxl.load_workbook(excel_file)
        sheet=excel[sheet_name]
        sheet.cell(row=row,column=col).value=data
        excel.save(excel_file)

